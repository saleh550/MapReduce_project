import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


# from openpyxl import calculation
def getSeason(argument):
    if argument == 'C':
        return "winter"
    if argument == 'F':
        return "spring"
    if argument == 'I':
        return "summer"
    if argument == 'L':
        return "fall"


def getMonth(argument):
    if argument == 'B':
        return 'jan'
    if argument == 'C':
        return 'feb'
    if argument == 'D':
        return 'march'
    if argument == 'E':
        return 'april'
    if argument == 'F':
        return 'may'
    if argument == 'G':
        return 'jun'
    if argument == 'H':
        return 'jul'
    if argument == 'I':
        return 'aug'
    if argument == 'J':
        return 'sep'
    if argument == 'K':
        return 'oct'
    if argument == 'L':
        return 'nov'
    if argument == 'M':
        return 'dec'


wb = load_workbook('./DataFile.xlsx', data_only=True)
ws = wb.active

max_investors_number_inYear = {
    "year": '',
    "investors": 0,
}
min_investors_number_inYear = {
    "year": '',
    "investors": 0,
}
max_investors_number_inMonth = {
    "month": '',
    "year": '',
    "investors": 0

}
min_investors_number_inMonth = {
    "month": '',
    "year": '',
    "investors": 0

}
max_investors_number_inSeason = {
    "season": '',
    "year": '',
    "investors": 0
}
min_investors_number_inSeason = {
    "season": '',
    "year": '',
    "investors": 0
}
Multi_year_average = 0
Multi_year_sum = 0
min_number_inYear = 350 * 12
max_number_inYear = 0
min_number_inMonth = 350
max_number_inMonth = 0
row_sum = 0
my_list = []
for row in range(1, 122):
    for col in range(1, 14):
        char = get_column_letter(col)
        if char != 'A':
            row_sum += ws[char + str(row)].value
            # sum of all investors
            Multi_year_sum += ws[char + str(row)].value
            # calculate the  maximum investors in month
            if ws[char + str(row)].value > max_number_inMonth:
                max_number_inMonth = ws[char + str(row)].value
                max_investors_number_inMonth['year'] = ws['A'+str(row)].value
                max_investors_number_inMonth['month'] = getMonth(char)
                max_investors_number_inMonth['investors'] = max_number_inMonth
            # calculate the  minimum investors in month
            if ws[char + str(row)].value < min_number_inMonth:
                min_number_inMonth = ws[char + str(row)].value
                min_investors_number_inMonth['year'] = ws['A'+str(row)].value
                min_investors_number_inMonth['month'] = getMonth(char)
                min_investors_number_inMonth['investors'] = min_number_inMonth

        if char == 'M':
            # calculate the max sums of rows for the maximum investors in year
            if row_sum > max_number_inYear:
                max_number_inYear = row_sum
                max_investors_number_inYear['year'] = ws['A'+str(row)].value
                max_investors_number_inYear['investors'] = max_number_inYear
                # calculate the min sums of rows for the minimum investors in year
            if row_sum < min_number_inYear:
                min_number_inYear = row_sum
                min_investors_number_inYear['year'] = ws['A'+str(row)].value
                min_investors_number_inYear['investors'] = min_number_inYear
            row_sum = 0

max_season_investors = 0
min_season_investors = 3*350
season_sum = 0
# This  loop goes through the months in the following way : 12 1 2 3 ... 11
for row in range(1, 122):
    for i in range(13, 26):
        col = i % 13
        if col == 0:
            col = 13
        if col == 1:
            continue
        char = get_column_letter(col)
        season_sum += ws[char + str(row)].value
        if char == 'C' or char == 'F' or char == 'I' or char == 'L':
            if season_sum > max_season_investors:
                max_season_investors = season_sum
                max_investors_number_inSeason['year'] = ws['A' + str(row)].value
                max_investors_number_inSeason['season'] = getSeason(char)
                max_investors_number_inSeason['investors'] = season_sum

            if season_sum < min_season_investors:
                min_season_investors = season_sum
                min_investors_number_inSeason['year'] = ws['A' + str(row)].value
                min_investors_number_inSeason['season'] = getSeason(char)
                min_investors_number_inSeason['investors'] = season_sum
            season_sum = 0


Multi_year_average = Multi_year_sum / 121


print("The multi year average is: " + str(Multi_year_average))
print("In " + str(max_investors_number_inYear['year']) + " was the max of investors of such " + str(max_investors_number_inYear['investors']))
print("In " + str(min_investors_number_inYear['year']) + " was the min of investors of such " + str(min_investors_number_inYear['investors']))
print("In " + str(max_investors_number_inMonth['month']) + " " + str(max_investors_number_inMonth['year']) + " was the maximum of investors of such " + str(max_investors_number_inMonth['investors']))
print("In " + str(min_investors_number_inMonth['month']) + " " + str(min_investors_number_inMonth['year']) + " was the minimum of investors of such " + str(min_investors_number_inMonth['investors']))
print("In " + str(max_investors_number_inSeason['season']) + " " + str(max_investors_number_inSeason['year']) + " was the maximum of investors of such " + str(max_investors_number_inSeason['investors']))
print("In " + str(min_investors_number_inSeason['season']) + " " + str(min_investors_number_inSeason['year']) + " was the minimum of investors of such " + str(min_investors_number_inSeason['investors']))
